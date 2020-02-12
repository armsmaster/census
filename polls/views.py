import datetime
import openpyxl
from django.shortcuts import render
from django.urls import reverse
from django.views import generic
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.template import loader
from django.http import HttpResponse, HttpResponseRedirect
from . import models
from . import forms


class ChoiceList_List(LoginRequiredMixin, generic.ListView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.ChoiceList
    template_name = 'polls/choice_list_list.html'
    context_object_name = 'items'


class ChoiceList_Detail(LoginRequiredMixin, generic.DetailView, generic.FormView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.ChoiceList
    template_name = 'polls/choice_list_detail.html'
    form_class = forms.Choice_Form

    def get_success_url(self):
        object = self.get_object()
        return reverse('polls:choice-list-detail', kwargs={'pk': object.pk})

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        choice = form.save(commit=False)
        choice.choice_list = self.get_object()
        choice.save()
        return super(ChoiceList_Detail, self).form_valid(form)


class ChoiceList_Create(LoginRequiredMixin, generic.edit.CreateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.ChoiceList
    form_class = forms.ChoiceList
    template_name = 'polls/choice_list_create.html'

    def get_success_url(self):
        return reverse('polls:choice-list-detail', kwargs={'pk': self.object.pk})


class ChoiceList_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.ChoiceList
    form_class = forms.ChoiceList
    template_name = 'polls/choice_list_update.html'

    def get_success_url(self):
        return reverse('polls:choice-list-detail', kwargs={'pk': self.object.pk})


class ChoiceList_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.ChoiceList
    template_name = 'polls/choice_list_confirm_delete.html'
    
    def delete(self, *args, **kwargs):
        self.object = self.get_object()
        for q in self.object.questions.all():
            q.choice_list = None
            q.save()
        return super(ChoiceList_Delete, self).delete(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('polls:choice-list-list')


class Choice_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Choice
    # fields = ['name', 'num_value']
    form_class = forms.Choice_Form
    template_name = 'polls/choice_update.html'

    def get_success_url(self):
        return reverse('polls:choice-list-detail', kwargs={'pk': self.object.choice_list.pk})


class Choice_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Choice
    template_name = 'polls/choice_confirm_delete.html'

    def get_success_url(self):
        return reverse('polls:choice-list-detail', kwargs={'pk': self.object.choice_list.pk})


class QuestionGroup_Create(LoginRequiredMixin, generic.edit.CreateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.QuestionGroup
    form_class = forms.QGroup
    template_name = 'polls/qgroup_create.html'

    def get_success_url(self):
        return reverse('polls:question-list')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lvl3 = [x.pk for x in models.QuestionGroup.objects.all() if x.level() >= 3]
        context['form'].fields['parent'].queryset = models.QuestionGroup.objects.all().exclude(pk__in=lvl3)
        return context


class QuestionGroup_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.QuestionGroup
    form_class = forms.QGroup
    template_name = 'polls/qgroup_update.html'

    def get_success_url(self):
        return reverse('polls:question-list')
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lvl3 = [x.pk for x in models.QuestionGroup.objects.all() if x.level() >= 3]
        context['form'].fields['parent'].queryset = models.QuestionGroup.objects.all().exclude(pk=context['object'].pk).exclude(pk__in=lvl3)
        return context


class QuestionGroup_Detail(LoginRequiredMixin, generic.DetailView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.QuestionGroup
    template_name = 'polls/qgroup_detail.html'


class QuestionGroup_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.QuestionGroup
    template_name = 'polls/qgroup_confirm_delete.html'
    
    def delete(self, *args, **kwargs):
        self.object = self.get_object()
        parent = self.object.parent
        for s in self.object.subgroups.all():
            s.parent = parent
            s.save()
        for q in self.object.questions.all():
            q.group = parent
            q.save()
        return super(QuestionGroup_Delete, self).delete(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('polls:question-list')


class Question_List(LoginRequiredMixin, generic.ListView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    template_name = 'polls/question_list.html'
    context_object_name = 'items'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['question_groups'] = sorted([x for x in models.QuestionGroup.objects.all()], key=lambda x: x.path())
        return context


class Question_Detail(LoginRequiredMixin, generic.DetailView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    template_name = 'polls/question_detail.html'


class Question_Create(LoginRequiredMixin, generic.edit.CreateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    # fields = ['name', 'text', 'data_type']
    form_class = forms.Question
    template_name = 'polls/question_create.html'

    def get_success_url(self):
        return reverse('polls:question-detail', kwargs={'pk': self.object.pk})


class Question_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    # fields = ['name', 'text', 'data_type']
    form_class = forms.Question
    template_name = 'polls/question_update.html'

    def get_success_url(self):
        return reverse('polls:question-detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        if form.instance.data_type in [models.Question.MSMC, models.Question.SSMC]:
            form.instance.range_min = None
            form.instance.range_max = None
            form.instance.range_step = None
        elif form.instance.data_type == models.Question.NUMBER_RANGE:
            form.instance.choice_list = None
        elif form.instance.data_type in [models.Question.TEXT, models.Question.NUMBER]:
            form.instance.range_min = None
            form.instance.range_max = None
            form.instance.range_step = None
            form.instance.choice_list = None
        return super().form_valid(form)


class Question_ChoiceList_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    # fields = ['choice_list']
    form_class = forms.Question_ChoiceList
    template_name = 'polls/question_update.html'

    def get_success_url(self):
        return reverse('polls:question-detail', kwargs={'pk': self.object.pk})


class Question_Range_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    # fields = ['range_min', 'range_max', 'range_step']
    form_class = forms.Question_Range
    template_name = 'polls/question_update.html'

    def get_success_url(self):
        return reverse('polls:question-detail', kwargs={'pk': self.object.pk})


class Question_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Question
    template_name = 'polls/question_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('polls:question-list')


class Survey_List(LoginRequiredMixin, generic.ListView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    template_name = 'polls/survey_list.html'
    context_object_name = 'items'


class Survey_Detail(LoginRequiredMixin, generic.DetailView, generic.FormView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    template_name = 'polls/survey_detail.html'
    form_class = forms.MapSurveyQuestion_Form

    def get_success_url(self):
        object = self.get_object()
        return '/survey-details/{}/'.format(object.id)
    
    def post(self, request, *args, **kwargs):
        # self.object = self.get_object()
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        msq = form.save(commit=False)
        msq.survey = self.get_object()
        msq.save()
        msq.tech_sort_order()
        return super(Survey_Detail, self).form_valid(form)


class Survey_Detail_Add_Person(LoginRequiredMixin, generic.DetailView, generic.FormView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    template_name = 'polls/survey_detail_add_person.html'
    form_class = forms.MapUserSurvey_Form

    def get_success_url(self):
        object = self.get_object()
        return '/survey-details/{}/'.format(object.id)
    
    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        msp = form.save(commit=False)
        msp.survey = self.get_object()
        msp.save()
        return super(Survey_Detail_Add_Person, self).form_valid(form)


class Survey_Create(LoginRequiredMixin, generic.edit.CreateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    # fields = ['name', 'description']
    form_class = forms.Survey
    template_name = 'polls/survey_create.html'

    def get_success_url(self):
        return reverse('polls:survey-detail', kwargs={'pk': self.object.pk})


class Survey_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    # fields = ['name', 'description']
    form_class = forms.Survey
    template_name = 'polls/survey_update.html'

    def get_success_url(self):
        return reverse('polls:survey-detail', kwargs={'pk': self.object.pk})


class Survey_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Survey
    template_name = 'polls/survey_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('polls:survey-list')


class Survey_Question_Detail(LoginRequiredMixin, generic.DetailView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapSurveyQuestion
    template_name = 'polls/survey_question_detail.html'


class Survey_Question_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapSurveyQuestion
    # fields = ['condition_question', 'condition_answer',]
    form_class = forms.MapSurveyQuestion_Condition_Form
    template_name = 'polls/survey_question_update.html'

    def get_success_url(self):
        return reverse('polls:survey-question-detail', kwargs={'pk': self.object.pk})


class Survey_Question_Set_Sort_Order(LoginRequiredMixin, generic.DetailView, generic.FormView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapSurveyQuestion
    template_name = 'polls/survey_question_sort_order.html'
    form_class = forms.MapSurveyQuestionSortOrder
    
    def get_success_url(self):
        return reverse('polls:survey-detail', kwargs={'pk': self.get_object().survey.pk})
    
    def get_initial(self):
        initial = super().get_initial()
        initial['obj'] = self.get_object()
        return initial
    
    def form_valid(self, form):
        # print(form)
        # print(form.cleaned_data['data'])
        # print(self.get_object())
        insert_after = int(form.cleaned_data['data'])
        obj = self.get_object()
        obj.sort_order_put_after(insert_after)
        # print('insert_after', insert_after)
        return super(Survey_Question_Set_Sort_Order, self).form_valid(form)


class Survey_Question_Update_Mandatory(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapSurveyQuestion
    # fields = ['is_mandatory',]
    form_class = forms.MapSurveyQuestion_Mandatory_Form
    template_name = 'polls/survey_question_update.html'

    def get_success_url(self):
        return reverse('polls:survey-question-detail', kwargs={'pk': self.object.pk})


class Survey_Question_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapSurveyQuestion
    template_name = 'polls/survey_question_confirm_delete.html'
    
    def delete(self, *args, **kwargs):
        self.object = self.get_object()
        ans = self.object.answers()
        if ans:
            ans.delete()
        return super(Survey_Question_Delete, self).delete(*args, **kwargs)
    
    def get_success_url(self):
        return reverse('polls:survey-detail', kwargs={'pk': self.object.survey.pk})


class Survey_Person_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.MapUserSurvey
    template_name = 'polls/survey_person_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('polls:survey-detail', kwargs={'pk': self.object.survey.pk})


class Person_List(LoginRequiredMixin, generic.ListView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Person
    template_name = 'polls/person_list.html'
    context_object_name = 'items'


class Person_Detail(LoginRequiredMixin, generic.DetailView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Person
    template_name = 'polls/person_detail.html'


class Person_Create(LoginRequiredMixin, generic.edit.CreateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Person
    # fields = ['email', 'name_first', 'name_second', 'name_last', 'birth_date', 'sex', 'title']
    form_class = forms.Person
    template_name = 'polls/person_create.html'

    def get_success_url(self):
        return reverse('polls:person-detail', kwargs={'pk': self.object.pk})


class Person_Update(LoginRequiredMixin, generic.UpdateView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Person
    # fields = ['email', 'name_first', 'name_second', 'name_last', 'birth_date', 'sex', 'title']
    form_class = forms.Person
    template_name = 'polls/person_update.html'

    def get_success_url(self):
        return reverse('polls:person-detail', kwargs={'pk': self.object.pk})


class Person_Delete(LoginRequiredMixin, generic.DeleteView):
    login_url = '/login/'
    redirect_field_name = 'next'
    model = models.Person
    template_name = 'polls/person_confirm_delete.html'
    
    def get_success_url(self):
        return reverse('polls:person-list')


def survey_instance(request, **kwargs):
    template = loader.get_template('polls/survey_start.html')
    survey_id = kwargs.get('id', None)
    secret_code = kwargs.get('secret', None)

    if not survey_id:
        return

    if not secret_code:
        return

    surv = models.MapUserSurvey.objects.filter(pk=survey_id).first()

    if not surv:
        return

    if surv.random_letters != secret_code:
        return

    if not surv.time_visit:
        surv.time_visit = datetime.datetime.now()
        surv.save()

    all_q = surv.survey.questions.all().order_by('id')
    if not all_q:
        next_q = -1
    else:
        next_q = all_q[0]

    context = {
        'object': surv,
        'next_q': next_q,
    }
    return HttpResponse(template.render(context, request))


def survey_step(request, **kwargs):
    template = loader.get_template('polls/survey_step.html')
    survey_id = kwargs.get('id', None)
    secret_code = kwargs.get('secret', None)
    q = kwargs.get('q', None)

    if not survey_id:
        return

    if not secret_code:
        return

    if not q:
        return

    surv = models.MapUserSurvey.objects.filter(pk=survey_id).first()

    if not surv:
        return

    if surv.random_letters != secret_code:
        return

    all_q = surv.survey.questions.all().order_by('id')
    if not all_q:
        return

    q_current = all_q.filter(id=q).first()

    if not q_current:
        return

    if not surv.time_start:
        surv.time_start = datetime.datetime.now()
        surv.save()

    all_q_next = all_q.filter(id__gt=q).order_by('id')

    if not all_q_next:
        next_q = None
    else:
        next_q = all_q_next[0]
    
    if q_current.has_condition():
        cond_q = q_current.condition_question
        cond_a = q_current.condition_answer
        cond_ok = True
        existing_a = models.Answer.objects.filter(survey_instance=surv, question=cond_q.question).first()
        if not existing_a:
            cond_ok = False
        else:
            if existing_a.data != cond_a:
                cond_ok = False
        if not cond_ok:
            if next_q:
                return HttpResponseRedirect(reverse('polls:survey-step', kwargs={'id': surv.pk, 'secret': surv.random_letters, 'q': next_q.pk}))
            else:
                return HttpResponseRedirect(reverse('polls:survey-end', kwargs={'id': surv.pk, 'secret': surv.random_letters}))
        
    
    if request.method == "POST":
        ok_to_go = False
        data_result = None
        data_result_num = None
        if q_current.question.data_type == models.Question.MSMC:
            choices = [(c.id, c.name) for c in q_current.question.choice_list.choices.all()]
            choices_dict = { str(c.id): c.name for c in q_current.question.choice_list.choices.all()}
            choices_dict_num = { str(c.id): c.num_value for c in q_current.question.choice_list.choices.all()}
            form = forms.MSMC(choices, request.POST)
            # form.fields['data'].choices = choices
            # print('xxxxxx MSMC')
            if form.is_valid():
                # print('MSMC', form.cleaned_data['data'])
                if form.cleaned_data['data'] == []:
                    if q_current.is_mandatory:
                        pass
                    else:
                        ok_to_go = True
                else:
                    ok_to_go = True
                    x = '|||'.join([choices_dict[i] for i in form.cleaned_data['data']])
                    data_result = x
                    x_num = ';'.join([str(choices_dict_num[i]) for i in form.cleaned_data['data']])
                    data_result_num = x_num
            
        if q_current.question.data_type == models.Question.SSMC:
            choices = [(c.id, c.name) for c in q_current.question.choice_list.choices.all()]
            choices_dict = { str(c.id): c.name for c in q_current.question.choice_list.choices.all()}
            choices_dict_num = { str(c.id): c.num_value for c in q_current.question.choice_list.choices.all()}
            form = forms.SSMC(choices, request.POST)
            # print(form)
            # form.fields['data'].choices = choices_x
            # print(form.fields['data'].choices)
            # print(form.fields['data'])
            # print('xxxxxx SSMC')
            # j = form.is_valid()
            # print(j)
            # for k, v in form.fields['data'].choices:
            #     print(k, ',', v, ',', isinstance(v, (list, tuple)))
            if form.is_valid():
                # print('SSMC', form.cleaned_data['data'])
                if form.cleaned_data['data'] == '':
                    if q_current.is_mandatory:
                        pass
                    else:
                        ok_to_go = True
                else:
                    ok_to_go = True
                    data_result = choices_dict[form.cleaned_data['data']]
                    data_result_num = str(choices_dict_num[form.cleaned_data['data']])
        
        if q_current.question.data_type == models.Question.NUMBER_RANGE:
            form = forms.NumRange(request.POST)
            if form.is_valid():
                # print('NUMBER_RANGE', form.cleaned_data['data'])
                if form.cleaned_data['data'] is None:
                    if q_current.is_mandatory:
                        pass
                    else:
                        ok_to_go = True
                else:
                    ok_to_go = True
                    data_result = form.cleaned_data['data']
                    data_result_num = data_result
                
        if q_current.question.data_type == models.Question.NUMBER:
            form = forms.Numeric(request.POST)
            if form.is_valid():
                # print('NUMBER', form.cleaned_data['data'])
                if form.cleaned_data['data'] is None:
                    if q_current.is_mandatory:
                        pass
                    else:
                        ok_to_go = True
                else:
                    ok_to_go = True
                    data_result = form.cleaned_data['data']
                    data_result_num = data_result
        
        if q_current.question.data_type == models.Question.TEXT:
            form = forms.Txt(request.POST)
            if form.is_valid():
                # print('TEXT', form.cleaned_data['data'])
                if form.cleaned_data['data'] == '':
                    if q_current.is_mandatory:
                        pass
                    else:
                        ok_to_go = True
                else:
                    ok_to_go = True
                    data_result = form.cleaned_data['data']
        
        print('###')
        print('q_current.question.data_type =', q_current.question.data_type)
        print('ok_to_go =', ok_to_go)
        print('data_result =', data_result)
        print('###')
        
        if not ok_to_go:
            return HttpResponseRedirect(reverse('polls:survey-step', kwargs={'id': surv.pk, 'secret': surv.random_letters, 'q': q_current.pk}))
        
        a = models.Answer.objects.filter(survey_instance=surv, question=q_current.question).first()
        if not a:
            a = models.Answer(survey_instance=surv, question=q_current.question)
        a.data = data_result
        a.data_num = data_result_num
        a.save()
        
        if next_q:
            return HttpResponseRedirect(reverse('polls:survey-step', kwargs={'id': surv.pk, 'secret': surv.random_letters, 'q': next_q.pk}))
        else:
            return HttpResponseRedirect(reverse('polls:survey-end', kwargs={'id': surv.pk, 'secret': surv.random_letters}))
    
    form = None
    
    if q_current.question.data_type == models.Question.MSMC:
        choices = [(c.id, c.name) for c in q_current.question.choice_list.choices.all()]
        form = forms.MSMC(choices)
    if q_current.question.data_type == models.Question.SSMC:
        choices = [(c.id, c.name) for c in q_current.question.choice_list.choices.all()]
        form = forms.SSMC(choices)
    if q_current.question.data_type == models.Question.NUMBER_RANGE:
        form = forms.NumRange()
    if q_current.question.data_type == models.Question.NUMBER:
        form = forms.Numeric()
    if q_current.question.data_type == models.Question.TEXT:
        form = forms.Txt()
    
    context = {
        'object': surv,
        'q_current': q_current,
        'form': form,
        'next_q': next_q,
    }
    return HttpResponse(template.render(context, request))


def survey_end(request, **kwargs):
    template = loader.get_template('polls/survey_end.html')
    survey_id = kwargs.get('id', None)
    secret_code = kwargs.get('secret', None)

    if not survey_id:
        return

    if not secret_code:
        return

    surv = models.MapUserSurvey.objects.filter(pk=survey_id).first()

    if not surv:
        return

    if surv.random_letters != secret_code:
        return

    if not surv.time_end:
        surv.time_end = datetime.datetime.now()
        surv.save()

    context = {
        'object': surv,
    }
    return HttpResponse(template.render(context, request))


def question_delete(request, **kwargs):
    
    question_id = kwargs.get('pk', None)

    if not question_id:
        return HttpResponseRedirect(reverse('polls:question-list'))
    
    question = models.Question.objects.filter(pk=question_id).first()
    if not question:
        return HttpResponseRedirect(reverse('polls:question-list'))

    if not question.can_delete():
        return HttpResponseRedirect(reverse('polls:question-list'))
    else:
        return HttpResponseRedirect(reverse('polls:question-delete-confirm', kwargs={'pk': question.pk}))


def export_answers_to_xlsx(request, **kwargs):
    survey_id = kwargs.get('id', None)
    survey = models.Survey.objects.filter(id=survey_id).first()
    answers = survey.answers()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={s}-answers.xlsx'.format(s=survey.id)
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    columns = [
        'person_id',
        'person_email',
        'survey_id',
        'survey_name',
        'question_id',
        'question_name',
        'answer',
        'answer_num',
    ]
    
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    for a in answers:
        row_num += 1
        worksheet.cell(row=row_num, column=1).value = a.survey_instance.person.id
        worksheet.cell(row=row_num, column=2).value = a.survey_instance.person.email
        worksheet.cell(row=row_num, column=3).value = a.survey_instance.survey.id
        worksheet.cell(row=row_num, column=4).value = a.survey_instance.survey.name
        worksheet.cell(row=row_num, column=5).value = a.question.id
        worksheet.cell(row=row_num, column=6).value = a.question.name
        worksheet.cell(row=row_num, column=7).value = a.data
        worksheet.cell(row=row_num, column=8).value = a.data_num
    
    workbook.save(response)
    return response


def index_news(request, **kwargs):
    template = loader.get_template('polls/index.html')
    context = {
    }
    return HttpResponse(template.render(context, request))
    

def user_guide(request, **kwargs):
    template = loader.get_template('polls/user_guide.html')
    context = {
    }
    return HttpResponse(template.render(context, request))
